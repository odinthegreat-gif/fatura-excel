import streamlit as st
import pdfplumber
import re
import io
import copy
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
import anthropic

# ── Sayfa ayarları ──────────────────────────────────────────────
st.set_page_config(
    page_title="Fatura → Excel Dönüştürücü",
    page_icon="📊",
    layout="wide",
)

# ── CSS ─────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1e3a5f 0%, #2d6a9f 100%);
        padding: 20px 30px; border-radius: 12px; margin-bottom: 25px;
        color: white;
    }
    .main-header h1 { margin: 0; font-size: 1.8rem; }
    .main-header p  { margin: 5px 0 0; opacity: 0.85; font-size: 0.95rem; }
    .metric-card {
        background: #f8f9fa; border-left: 4px solid #2d6a9f;
        padding: 14px 18px; border-radius: 8px; margin: 6px 0;
    }
    .metric-card .label { font-size: 0.78rem; color: #666; font-weight: 600; text-transform: uppercase; }
    .metric-card .value { font-size: 1.25rem; font-weight: 700; color: #1e3a5f; }
    .success-box {
        background: #d4edda; border: 1px solid #c3e6cb;
        padding: 12px 16px; border-radius: 8px; color: #155724;
    }
    .warning-box {
        background: #fff3cd; border: 1px solid #ffc107;
        padding: 12px 16px; border-radius: 8px; color: #856404;
    }
    .stProgress > div > div { background-color: #2d6a9f !important; }
    div[data-testid="stFileUploader"] {
        border: 2px dashed #2d6a9f !important;
        border-radius: 10px !important; padding: 10px;
    }
</style>
""", unsafe_allow_html=True)

# ── Header ───────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <h1>📊 Fatura → Excel Dönüştürücü</h1>
  <p>PDF faturalarınızı yükleyin, yapay zeka otomatik olarak verileri okuyup Excel tablosuna aktarsın.</p>
</div>
""", unsafe_allow_html=True)

# ── Sabit veriler ────────────────────────────────────────────────
MONTHS_TR = {
    1:"OCAK", 2:"ŞUBAT", 3:"MART", 4:"NİSAN",
    5:"MAYIS", 6:"HAZİRAN", 7:"TEMMUZ", 8:"AĞUSTOS",
    9:"EYLÜL", 10:"EKİM", 11:"KASIM", 12:"ARALIK",
}
MONTHS_ROW = {v: 4+i for i, v in enumerate(MONTHS_TR.values())}  # OCAK→4 … ARALIK→15

SHEETS = {
    "METYX MNS2 ŞUBESİ": ["MNS2", "METYX MNS2"],
    "MANİSA ŞUBESİ":      ["MANİSA ŞUBESİ", "MANİSA SUBESI", "MUTFAK"],
}

# Excel sütun haritası (A=1 … M=13)
COL_MAP = {
    "elektrik_kwh":   4,   # D
    "elektrik_tl":    5,   # E
    "su_m3":          6,   # F
    "su_tl":          7,   # G
    "dogalgaz_kwh":   8,   # H
    "dogalgaz_tl":    9,   # I
    "atik_adet":     10,   # J
    "atik_tl":       11,   # K
    "atiksu_m3":     12,   # L
    "atiksu_tl":     13,   # M
}

# ── PDF okuma yardımcıları ───────────────────────────────────────
def extract_pdf_text(uploaded_file) -> str:
    """pdfplumber ile tüm sayfaları metin olarak döndürür."""
    pages = []
    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                pages.append(t)
    return "\n\n--- SAYFA SONU ---\n\n".join(pages)

# ── Claude ile fatura ayrıştırma ─────────────────────────────────
SYSTEM_PROMPT = """
Sen bir fatura veri çıkarma uzmanısın. Sana Türkçe OSB (Organize Sanayi Bölgesi) faturaları verilecek.
Her faturadan şu bilgileri JSON formatında çıkar:

{
  "faturalar": [
    {
      "tesis": "METYX MNS2 ŞUBESİ" veya "MANİSA ŞUBESİ",
      "ay": 1-12 arası sayı,
      "yil": 4 haneli yıl,
      "fatura_tipi": "elektrik" | "su" | "dogalgaz" | "atik" | "atiksu" | "izin_belgesi" | "altyapi" | "diger",
      "miktar": sayı veya null,
      "birim": "KWH" | "M3" | "ADET" veya null,
      "tutar_kdv_dahil": sayı (ödenecek tutar)
    }
  ]
}

Fatura tarihi: son ödeme tarihi değil, FATURA TARİHİ'nden ayı belirle.
Mutfak aboneliği MANİSA ŞUBESİ'ne aittir.
MEL fatura numarası = elektrik
MSU fatura numarası = su (şebeke suyu)
MDG fatura numarası = doğalgaz
MKT fatura numarası = evsel katı atık
MAT fatura numarası = atıksu (KOI, AKM, YAĞ&GRES içerir)
GNL/MTA fatura numarası = diğer hizmetler (izin belgesi, altyapı)

Çıktı SADECE JSON olsun, başka metin ekleme.
"""

def parse_invoices_with_claude(pdf_text: str) -> list[dict]:
    """Claude API ile faturaları ayrıştır."""
    client = anthropic.Anthropic()
    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4096,
        messages=[{
            "role": "user",
            "content": f"Aşağıdaki PDF fatura metinlerini analiz et:\n\n{pdf_text}"
        }],
        system=SYSTEM_PROMPT,
    )
    raw = msg.content[0].text.strip()
    # JSON kod bloğu varsa temizle
    raw = re.sub(r"^```json\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    data = json.loads(raw)
    return data.get("faturalar", [])

# ── Excel doldurma ───────────────────────────────────────────────
def fill_excel(template_bytes: bytes, invoices: list[dict]) -> bytes:
    """Şablonu doldurup dolu Excel döndürür."""
    wb = load_workbook(io.BytesIO(template_bytes))

    for inv in invoices:
        # Tesis → sheet
        sheet_name = None
        tesis_val = (inv.get("tesis") or "").upper()
        for sn, keywords in SHEETS.items():
            if any(k.upper() in tesis_val for k in keywords):
                sheet_name = sn
                break
        if sheet_name is None or sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        ay_num = inv.get("ay")
        if not ay_num:
            continue
        row = MONTHS_ROW.get(MONTHS_TR.get(ay_num, ""), None)
        if row is None:
            continue

        tip = (inv.get("fatura_tipi") or "").lower()
        miktar = inv.get("miktar")
        tutar = inv.get("tutar_kdv_dahil")

        if tip == "elektrik":
            if miktar:  ws.cell(row=row, column=COL_MAP["elektrik_kwh"]).value = miktar
            if tutar:   ws.cell(row=row, column=COL_MAP["elektrik_tl"]).value  = tutar
        elif tip == "su":
            if miktar:  ws.cell(row=row, column=COL_MAP["su_m3"]).value = miktar
            if tutar:   ws.cell(row=row, column=COL_MAP["su_tl"]).value  = tutar
        elif tip == "dogalgaz":
            prev_kwh = ws.cell(row=row, column=COL_MAP["dogalgaz_kwh"]).value or 0
            prev_tl  = ws.cell(row=row, column=COL_MAP["dogalgaz_tl"]).value  or 0
            if miktar:  ws.cell(row=row, column=COL_MAP["dogalgaz_kwh"]).value = prev_kwh + miktar
            if tutar:   ws.cell(row=row, column=COL_MAP["dogalgaz_tl"]).value  = prev_tl  + tutar
        elif tip == "atik":
            if miktar:  ws.cell(row=row, column=COL_MAP["atik_adet"]).value = miktar
            if tutar:   ws.cell(row=row, column=COL_MAP["atik_tl"]).value   = tutar
        elif tip == "atiksu":
            prev_m3 = ws.cell(row=row, column=COL_MAP["atiksu_m3"]).value or 0
            prev_tl = ws.cell(row=row, column=COL_MAP["atiksu_tl"]).value or 0
            if miktar:  ws.cell(row=row, column=COL_MAP["atiksu_m3"]).value = prev_m3 + miktar
            if tutar:   ws.cell(row=row, column=COL_MAP["atiksu_tl"]).value = prev_tl + tutar

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# ── Özet tablo ───────────────────────────────────────────────────
def build_summary(invoices: list[dict]) -> dict:
    summary = {}
    for inv in invoices:
        tesis = inv.get("tesis", "Bilinmeyen")
        ay    = MONTHS_TR.get(inv.get("ay", 0), "?")
        tip   = inv.get("fatura_tipi", "?")
        tutar = inv.get("tutar_kdv_dahil") or 0
        key   = (tesis, ay)
        if key not in summary:
            summary[key] = {}
        summary[key][tip] = summary[key].get(tip, 0) + tutar
    return summary

# ════════════════════════════════════════════════════════════════
# ── Streamlit UI ────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════

col_left, col_right = st.columns([1, 1], gap="large")

with col_left:
    st.subheader("📁 Dosya Yükleme")

    pdf_files = st.file_uploader(
        "PDF Faturalar (birden fazla seçebilirsiniz)",
        type=["pdf"],
        accept_multiple_files=True,
        key="pdf_upload",
    )

    xlsx_file = st.file_uploader(
        "Excel Şablonu (.xlsx)",
        type=["xlsx"],
        key="xlsx_upload",
        help="Doldurulacak Excel şablonunu yükleyin",
    )

    if pdf_files:
        st.markdown(f"<div class='success-box'>✅ {len(pdf_files)} PDF yüklendi</div>", unsafe_allow_html=True)
    if xlsx_file:
        st.markdown("<div class='success-box'>✅ Excel şablonu yüklendi</div>", unsafe_allow_html=True)

    st.markdown("---")

    if not pdf_files:
        st.info("👆 Lütfen PDF fatura dosyalarını yükleyin.")
    elif not xlsx_file:
        st.warning("⚠️ Excel şablonu yüklenmedi. Varsayılan şablon oluşturulacak.")

    process_btn = st.button(
        "🚀 Faturları İşle ve Excel Oluştur",
        type="primary",
        use_container_width=True,
        disabled=(not pdf_files),
    )

with col_right:
    st.subheader("📋 Nasıl Çalışır?")
    st.markdown("""
    **Adım 1:** PDF fatura dosyalarınızı yükleyin  
    **Adım 2:** Excel şablonunuzu yükleyin *(opsiyonel)*  
    **Adım 3:** "Faturları İşle" butonuna tıklayın  
    **Adım 4:** Doldurulmuş Excel dosyasını indirin  

    ---
    **Desteklenen Fatura Tipleri:**
    | Kod | Tür |
    |-----|-----|
    | MEL | Elektrik |
    | MSU | Su |
    | MDG | Doğalgaz |
    | MKT | Evsel Katı Atık |
    | MAT | Atıksu |
    | GNL/MTA | Diğer Hizmetler |

    **Desteklenen Tesisler:**
    - METYX MNS2 ŞUBESİ
    - MANİSA ŞUBESİ (Mutfak dahil)
    """)

# ── İşleme ──────────────────────────────────────────────────────
if process_btn and pdf_files:
    st.markdown("---")
    st.subheader("⚙️ İşleniyor...")

    progress = st.progress(0)
    status   = st.empty()

    all_invoices = []

    try:
        # 1. PDF metinleri çıkar
        status.info("📄 PDF dosyaları okunuyor...")
        combined_text = ""
        for i, pdf_file in enumerate(pdf_files):
            status.info(f"📄 Okunuyor: {pdf_file.name}")
            combined_text += extract_pdf_text(pdf_file) + "\n\n===\n\n"
            progress.progress(int((i + 1) / len(pdf_files) * 40))

        # 2. Claude ile ayrıştır
        status.info("🤖 Yapay zeka fatura verilerini analiz ediyor...")
        progress.progress(50)

        # Büyük metinleri böl (token sınırı)
        chunk_size = 80_000
        chunks = [combined_text[i:i+chunk_size] for i in range(0, len(combined_text), chunk_size)]

        for idx, chunk in enumerate(chunks):
            status.info(f"🤖 Analiz ediliyor... ({idx+1}/{len(chunks)})")
            invoices = parse_invoices_with_claude(chunk)
            all_invoices.extend(invoices)
            progress.progress(50 + int((idx + 1) / len(chunks) * 30))

        # 3. Excel doldur
        status.info("📊 Excel dosyası oluşturuluyor...")
        progress.progress(85)

        # Şablon al
        if xlsx_file:
            template_bytes = xlsx_file.read()
        else:
            # Minimal varsayılan şablon oluştur
            wb = openpyxl.Workbook()
            for sn in ["METYX MNS2 ŞUBESİ", "MANİSA ŞUBESİ"]:
                ws = wb.create_sheet(sn)
                ws['A1'] = ' FATURA VERİLERİ TABLOSU (2026 YILI)'
                headers = ['Sıra No','Ay','Çalışan Sayısı',
                           'Elektrik (kwh)','Fatura Bedeli (TL)',
                           'Su (m3)','Fatura Bedeli (TL)',
                           'Doğalgaz (kwh)','Fatura Bedeli (TL)',
                           'Evsel Katı Atık (Adet)','Fatura Bedeli (TL)',
                           'Atıksu (m3)','Fatura Bedeli (TL)']
                for j, h in enumerate(headers, 1):
                    ws.cell(row=2, column=j).value = h
                months = list(MONTHS_TR.values())
                for i, m in enumerate(months, 1):
                    ws.cell(row=3+i, column=1).value = i
                    ws.cell(row=3+i, column=2).value = m
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            buf = io.BytesIO()
            wb.save(buf)
            template_bytes = buf.getvalue()

        filled_bytes = fill_excel(template_bytes, all_invoices)
        progress.progress(100)

        # ── Sonuçlar ──────────────────────────────────────────────
        status.empty()
        st.success(f"✅ İşlem tamamlandı! **{len(all_invoices)}** fatura kaydı işlendi.")

        # Özet metrikler
        tip_counts = {}
        tesis_counts = {}
        for inv in all_invoices:
            t = inv.get("fatura_tipi", "?")
            ts = inv.get("tesis", "?")
            tip_counts[t]   = tip_counts.get(t, 0) + 1
            tesis_counts[ts] = tesis_counts.get(ts, 0) + 1

        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(f"""
            <div class='metric-card'>
              <div class='label'>Toplam Fatura</div>
              <div class='value'>{len(all_invoices)}</div>
            </div>""", unsafe_allow_html=True)
        with c2:
            st.markdown(f"""
            <div class='metric-card'>
              <div class='label'>PDF Sayısı</div>
              <div class='value'>{len(pdf_files)}</div>
            </div>""", unsafe_allow_html=True)
        with c3:
            total_tl = sum(inv.get("tutar_kdv_dahil") or 0 for inv in all_invoices)
            st.markdown(f"""
            <div class='metric-card'>
              <div class='label'>Toplam Tutar</div>
              <div class='value'>{total_tl:,.0f} TL</div>
            </div>""", unsafe_allow_html=True)

        # Detay tablosu
        if all_invoices:
            st.markdown("### 📋 Tespit Edilen Faturalar")
            tip_icons = {
                "elektrik":"⚡","su":"💧","dogalgaz":"🔥",
                "atik":"🗑️","atiksu":"🌊","izin_belgesi":"📜",
                "altyapi":"🏗️","diger":"📄",
            }
            rows = []
            for inv in all_invoices:
                icon = tip_icons.get(inv.get("fatura_tipi",""), "📄")
                rows.append({
                    "Tesis": inv.get("tesis","?"),
                    "Ay": MONTHS_TR.get(inv.get("ay",0),"?"),
                    "Yıl": inv.get("yil","?"),
                    "Tür": f"{icon} {inv.get('fatura_tipi','?').title()}",
                    "Miktar": f"{inv.get('miktar',''):,}" if inv.get("miktar") else "-",
                    "Birim": inv.get("birim") or "-",
                    "Tutar (TL)": f"{inv.get('tutar_kdv_dahil',0):,.2f}" if inv.get("tutar_kdv_dahil") else "-",
                })
            import pandas as pd
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        # İndir butonu
        st.download_button(
            label="⬇️ Excel Dosyasını İndir",
            data=filled_bytes,
            file_name="2026_Yili_Fatura_Verileri.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

    except json.JSONDecodeError as e:
        status.empty()
        progress.empty()
        st.error(f"❌ JSON ayrıştırma hatası: {e}")
        st.info("Claude API yanıtı beklenen formatta değil. Lütfen tekrar deneyin.")
    except Exception as e:
        status.empty()
        progress.empty()
        st.error(f"❌ Hata oluştu: {e}")
        import traceback
        with st.expander("Hata detayları"):
            st.code(traceback.format_exc())

# ── Alt bilgi ────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    "<p style='text-align:center;color:#888;font-size:0.82rem;'>"
    "MANİSA OSB · Fatura Takip Sistemi · 2026"
    "</p>",
    unsafe_allow_html=True,
)
